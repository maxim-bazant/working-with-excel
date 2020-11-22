# working with excel
import openpyxl as xl
from openpyxl.chart import BarChart, Reference      # with this we can add graph to our excel work


def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row):
        cell = sheet.cell(row, 3)                     # this is how we get into each cell
        corrected_price = cell.value * 0.9            # and change the value
        corrected_price_cell = sheet.cell(row, 4)     # we create new column
        corrected_price_cell.value = corrected_price  # here we add to the new cell(in the new column) the corrected price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,   # this is how we select more values
                       min_col=4,               # and set keyword arguments
                       max_col=4                # we will use these values to create graph later
                       )

    chart = BarChart()                       # chart = graph
    chart.add_data(values)
    sheet.add_chart(chart, "e2")             # (chart - the grapf we want to add, "e2" - where we want to add the graph)

    wb.save(filename)  # this is how we save it(to new file called transactions2.xlsx)


def sum(filename, column):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    total_sum = 0

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, column)
        total_sum += cell.value

    total_sum_cell = sheet.cell(sheet.max_row + 1, column)
    total_sum_cell.value = total_sum

    wb.save(filename)


sum("transactions - kopie.xlsx", column=3)
process_workbook("transactions2.xlsx")


